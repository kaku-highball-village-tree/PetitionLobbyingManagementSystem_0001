#!/usr/bin/env python3
"""PetitionLobbyingManagementSystem_Cmd.py

Excel を読み込み、TSV と 1相談=1TXT を生成する。
"""

from __future__ import annotations

from pathlib import Path
from datetime import datetime, date
import re
import sys

from openpyxl import load_workbook


LIST_REQUIRED_HEADERS = [
    "管理番号",
    "受付番号",
    "相談日",
    "相談者",
    "相談内容(表題)",
    "依頼日",
    "依頼課",
    "受付担当者",
    "処理可否_日",
    "処理の可否_1",
    "処理の可否_2",
    "本文",
]

RE_INVALID_FILENAME_CHARACTERS = re.compile(r"[\\/:*?\"<>|]")


def psz_to_text(objValue: object) -> str:
    if objValue is None:
        return ""
    if isinstance(objValue, datetime):
        return objValue.strftime("%Y年%m月%d日")
    if isinstance(objValue, date):
        return objValue.strftime("%Y年%m月%d日")
    return str(objValue).strip()


def psz_sanitize_file_name(pszSourceText: str) -> str:
    pszReplacedText = RE_INVALID_FILENAME_CHARACTERS.sub("_", pszSourceText)
    return pszReplacedText.strip()


def b_is_blank_row(listRowValues: list[str]) -> bool:
    if not listRowValues:
        return True
    if all(pszItem == "" for pszItem in listRowValues):
        return True
    if listRowValues[0] == "":
        return True
    return False


def obj_build_output_paths(pathExcelFile: Path) -> tuple[Path, Path, Path]:
    pathBaseDirectory = pathExcelFile.parent
    pathTsvDirectory = pathBaseDirectory / "TSV"
    pathTextDirectory = pathBaseDirectory / "Text"

    pathTsvDirectory.mkdir(parents=True, exist_ok=True)
    pathTextDirectory.mkdir(parents=True, exist_ok=True)

    pszFileStem = pathExcelFile.stem
    pathTsvFile = pathTsvDirectory / f"{pszFileStem}_search.tsv"
    return pathTsvFile, pathTsvDirectory, pathTextDirectory


def i_main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python PetitionLobbyingManagementSystem_Cmd.py <excel_file.xlsx>")
        return 1

    pathExcelFile = Path(sys.argv[1]).expanduser().resolve()
    if not pathExcelFile.exists() or pathExcelFile.suffix.lower() != ".xlsx":
        print(f"ERROR: Excelファイルが見つからないか、xlsxではありません: {pathExcelFile}")
        return 1

    try:
        objWorkbook = load_workbook(filename=pathExcelFile, data_only=True)
    except Exception as objException:
        print(f"ERROR: Excel読込に失敗しました: {objException}")
        return 1

    if "相談一覧" not in objWorkbook.sheetnames or "依頼課" not in objWorkbook.sheetnames:
        print("ERROR: 必須シート(相談一覧 / 依頼課) が存在しません")
        return 1

    objSheet = objWorkbook["相談一覧"]
    listHeaderRow = [psz_to_text(objCell.value) for objCell in objSheet[1][: len(LIST_REQUIRED_HEADERS)]]
    if listHeaderRow != LIST_REQUIRED_HEADERS:
        print("ERROR: ヘッダー行が仕様と一致しません")
        print("EXPECTED:", "\t".join(LIST_REQUIRED_HEADERS))
        print("ACTUAL  :", "\t".join(listHeaderRow))
        return 1

    pathTsvFile, _pathTsvDirectory, pathTextDirectory = obj_build_output_paths(pathExcelFile)

    listTsvLines: list[str] = ["\t".join(LIST_REQUIRED_HEADERS)]
    iProcessedCount = 0
    iSkippedCount = 0

    for objRow in objSheet.iter_rows(min_row=2, max_col=len(LIST_REQUIRED_HEADERS)):
        listRowValues = [psz_to_text(objCell.value) for objCell in objRow]

        if b_is_blank_row(listRowValues):
            iSkippedCount += 1
            continue

        listTsvLines.append("\t".join(listRowValues))

        pszManagementNumber = psz_sanitize_file_name(listRowValues[0])
        pszReceptionNumber = psz_sanitize_file_name(listRowValues[1])
        pszTitleText = psz_sanitize_file_name(listRowValues[4])

        if pszTitleText == "":
            pszTitleText = "無題"

        pszTitleText = pszTitleText[:50]

        pszTextFileName = f"{pszManagementNumber}_{pszReceptionNumber}_{pszTitleText}.txt"
        pathTextFile = pathTextDirectory / pszTextFileName

        pszTextBody = (
            f"管理番号: {listRowValues[0]}\r\n"
            f"受付番号: {listRowValues[1]}\r\n"
            f"相談日: {listRowValues[2]}\r\n"
            f"相談者: {listRowValues[3]}\r\n"
            f"相談内容(表題): {listRowValues[4]}\r\n"
            f"依頼日: {listRowValues[5]}\r\n"
            f"依頼課: {listRowValues[6]}\r\n"
            f"受付担当者: {listRowValues[7]}\r\n"
            f"処理可否_日: {listRowValues[8]}\r\n"
            f"処理の可否_1: {listRowValues[9]}\r\n"
            f"処理の可否_2: {listRowValues[10]}\r\n"
            "\r\n"
            f"本文:\r\n{listRowValues[11]}\r\n"
        )

        try:
            pathTextFile.write_text(pszTextBody, encoding="utf-8", newline="")
            iProcessedCount += 1
        except Exception as objException:
            print(f"WARNING: TXT出力をスキップしました: {pathTextFile} ({objException})")

    try:
        pathTsvFile.write_text("\r\n".join(listTsvLines) + "\r\n", encoding="utf-8", newline="")
    except Exception as objException:
        print(f"ERROR: TSV出力に失敗しました: {objException}")
        return 1

    print("OK: 変換が完了しました")
    print(f"TSV: {pathTsvFile}")
    print(f"TXT件数: {iProcessedCount}")
    print(f"スキップ行: {iSkippedCount}")
    return 0


if __name__ == "__main__":
    raise SystemExit(i_main())
